"""
attendance_scraper.py
======================
Scrapes the Attendance System, handles React dropdowns, filters out
zero-attendance/undefined rows, computes overall stats using two methods,
and saves to a formatted Excel file.
"""

import os
import re
import sys
import time

import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException
from webdriver_manager.chrome import ChromeDriverManager

# --------------------------------------------------------------------------
# CONFIG -- edit these before running
# --------------------------------------------------------------------------
BASE_URL = "https://attendence-system-1910.vercel.app"
LOGIN_URL = f"{BASE_URL}/users/login"

HEADLESS = False         # keep False for your first run so you can watch it work
WAIT_TIMEOUT = 20        # seconds to wait for each page/element to appear
OUTPUT_FILE = "attendance_report.xlsx"

COURSE = None
BATCH = None
DIVISION = None
SEMESTER = None
# --------------------------------------------------------------------------

def get_credentials():
    """Read credentials from environment variables."""
    email = os.getenv("ATTENDANCE_EMAIL")
    password = os.getenv("ATTENDANCE_PASSWORD")

    if not email or not password:
        raise RuntimeError(
            "Missing credentials. Set ATTENDANCE_EMAIL and ATTENDANCE_PASSWORD."
        )

    return email, password

def build_driver():
    options = Options()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--incognito")
    options.add_argument("--window-size=1440,1000")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def xpath_quote(s):
    if "'" not in s:
        return f"'{s}'"
    parts = s.split("'")
    return "concat('" + "', \"'\", '".join(parts) + "')"

def find_button(driver, text):
    return driver.find_element(By.XPATH, f"//button[normalize-space()={xpath_quote(text)}]")

def safe_click(driver, element):
    """Fallback click if normal click is intercepted by overlays or React renders."""
    try:
        driver.execute_script("arguments[0].scrollIntoView(true);", element)
        element.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", element)

def set_dropdown_if_specified(driver, wait, label_text, value):
    """React dropdown handler for setting a specific value."""
    if value is None:
        return

    xpath_label = f"//*[contains(text(), {xpath_quote(label_text)})]"
    container = wait.until(EC.presence_of_element_located(
        (By.XPATH, f"{xpath_label}/following-sibling::div | {xpath_label}/parent::*/following-sibling::div")
    ))
    safe_click(driver, container)
    time.sleep(0.5)

    option = wait.until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                f"//*[normalize-space()={xpath_quote(value)} "
                f"and not(ancestor-or-self::*[@aria-hidden='true'])]"
            )
        )
    )

    safe_click(driver, option)
    time.sleep(0.6)

def get_subject_names(driver, wait):
    """React dropdown handler to scrape all available subjects."""
    xpath_label = "//*[contains(text(), 'Select Subjects')]"
    container = wait.until(EC.presence_of_element_located(
        (By.XPATH, f"{xpath_label}/following-sibling::div | {xpath_label}/parent::*/following-sibling::div")
    ))
    safe_click(driver, container)
    time.sleep(1)

    options = driver.find_elements(By.XPATH, "//li | //*[@role='option']")

    subjects = []
    for opt in options:
        text = opt.text.strip()
        if text and text.lower() != "none" and text not in subjects:
            subjects.append(text)

    # Close the React dropdown by clicking the selector again.
    # Using ESC can leave some React dropdown implementations in a
    # state where the next Selenium click does not open the menu correctly.
    safe_click(driver, container)
    time.sleep(0.5)

    return subjects

def login(driver, wait, email, password):
    driver.get(LOGIN_URL)
    email_input = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder='john@gmail.com'], input[type='email']"))
    )
    password_input = driver.find_element(By.CSS_SELECTOR, "input[type='password']")

    email_input.clear()
    email_input.send_keys(email)
    password_input.clear()
    password_input.send_keys(password)

    find_button(driver, "Log In").click()

    try:
        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//button[normalize-space()='Your Attendances']")
        ))
    except TimeoutException:
        raise RuntimeError("Login failed. Check credentials or site status.")

def open_subject_selector(driver, wait):
    find_button(driver, "Your Attendances").click()
    wait.until(EC.presence_of_element_located(
        (By.XPATH, "//*[contains(text(),'Select Subject For Attendance')]")
    ))
    time.sleep(3)

def extract_summary(body_text):
    patterns = {
        "Subject": r"Subject:\s*(.+)",
        "Total Attendances": r"Total Attendances:\s*([a-zA-Z0-9]+)",
        "Total Present": r"Total Present:\s*([a-zA-Z0-9]+)",
        "Percentage": r"Your Percentages?:\s*([a-zA-Z0-9.%]+)",
    }
    row = {}
    for key, pattern in patterns.items():
        m = re.search(pattern, body_text)
        row[key] = m.group(1).strip() if m else None
    return row

def scrape_one_subject(driver, wait, subject_name):
    set_dropdown_if_specified(driver, wait, "Select Course", COURSE)
    set_dropdown_if_specified(driver, wait, "Select Batch", BATCH)
    set_dropdown_if_specified(driver, wait, "Select Division", DIVISION)
    set_dropdown_if_specified(driver, wait, "Select Semester", SEMESTER)

    set_dropdown_if_specified(driver, wait, "Select Subjects", subject_name)

    find_button(driver, "View Attendance").click()

    # Wait until the page loads the specific subject's name in the text
    wait.until(lambda d: subject_name.lower() in d.find_element(By.TAG_NAME, "body").text.lower())

    body_text = driver.find_element(By.TAG_NAME, "body").text
    row = extract_summary(body_text)

    find_button(driver, "Go Back").click()
    wait.until(EC.presence_of_element_located(
        (By.XPATH, "//*[contains(text(),'Select Subject For Attendance')]")
    ))
    time.sleep(1) # stabilization pause
    return row

def build_excel(rows, path):
    # Initialize DataFrame with scraped data
    df = pd.DataFrame(rows, columns=["Subject", "Total Attendances", "Total Present", "Percentage"])

    # Convert data types to string temporarily for cleanup processing
    for col in df.columns:
        df[col] = df[col].astype(str)

    # 1. Clean data: Drop undefined or missing rows
    df = df[~df["Subject"].str.lower().isin(["undefined", "none", "nan", ""])]
    df = df[~df["Total Attendances"].str.lower().isin(["undefined", "none", "nan", ""])]
    df = df[~df["Total Present"].str.lower().isin(["undefined", "none", "nan", ""])]

    # 2. Convert to numeric for calculations
    df["Total Attendances_num"] = pd.to_numeric(df["Total Attendances"], errors='coerce')
    df["Total Present_num"] = pd.to_numeric(df["Total Present"], errors='coerce')

    df = df.dropna(subset=["Total Attendances_num", "Total Present_num"])

    # 3. Filter out rows where Total Attendances or Total Present is 0
    df = df[(df["Total Attendances_num"] > 0) & (df["Total Present_num"] > 0)].copy()

    if df.empty:
        print("\nAll subjects were filtered out (undefined or 0 attendance). No Excel file generated.")
        return

    # Keep only the original target columns and convert to proper types
    df = df[["Subject", "Total Attendances", "Total Present", "Percentage"]].copy()
    df["Total Attendances"] = df["Total Attendances"].astype(int)
    df["Total Present"] = df["Total Present"].astype(int)

    # Save to Excel with proper formatting
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")
        ws = writer.book["Attendance"]

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add auto filter
        ws.auto_filter.ref = ws.dimensions

        # Set column widths
        widths = {"A": 48, "B": 20, "C": 18, "D": 15}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    # Now apply styling and add totals row using openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb["Attendance"]
    last_row = ws.max_row
    total_row = last_row + 2

    # Add TOTAL/AVERAGE row
    ws.cell(total_row, 1, "TOTAL / AVERAGE")
    ws.cell(total_row, 2, f"=SUM(B2:B{last_row})")
    ws.cell(total_row, 3, f"=SUM(C2:C{last_row})")
    ws.cell(total_row, 4, f"=AVERAGE(D2:D{last_row})")

    # Convert percentage strings to numeric percentages so Excel can average them correctly
    for row in range(2, last_row + 1):
        cell = ws.cell(row, 4)
        if isinstance(cell.value, str) and cell.value.endswith("%"):
            cell.value = float(cell.value[:-1]) / 100
            cell.number_format = "0.00%"

    # Format header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Format data rows with alternating colors and borders
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row_idx in range(2, last_row + 1):
        for col_idx in range(1, 5):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border

            # Alternate row colors
            if (row_idx - 2) % 2 == 1:
                cell.fill = alt_row_fill

            # Apply alignment based on column
            if col_idx == 1:  # Subject
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [2, 3]:  # Numbers
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 4:  # Percentage
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Format total row
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    total_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, 5):
        cell = ws.cell(total_row, col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border

        if col_idx == 1:  # Subject label
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = total_alignment

        # Apply percentage format to average cell
        if col_idx == 4:
            cell.number_format = "0.00%"

    wb.save(path)
    print(f"\nFiltered and saved successfully to {os.path.abspath(path)}")

def main():
    email, password = get_credentials()
    driver = build_driver()
    wait = WebDriverWait(driver, WAIT_TIMEOUT)
    results = []

    try:
        print("Logging in...")
        login(driver, wait, email, password)

        print("Opening subject selector...")
        open_subject_selector(driver, wait)

        set_dropdown_if_specified(driver, wait, "Select Course", COURSE)
        set_dropdown_if_specified(driver, wait, "Select Batch", BATCH)
        set_dropdown_if_specified(driver, wait, "Select Division", DIVISION)
        set_dropdown_if_specified(driver, wait, "Select Semester", SEMESTER)

        subjects = get_subject_names(driver, wait)
        print(f"Found {len(subjects)} subject(s): {', '.join(subjects)}")

        for subject in subjects:
            print(f"  Scraping '{subject}'...")
            for attempt in range(2):
                try:
                    row = scrape_one_subject(driver, wait, subject)
                    results.append(row)
                    print(f"    -> {row}")
                    break
                except StaleElementReferenceException:
                    if attempt == 1:
                        raise
                    continue
                except Exception as e:
                    print(f"    !! Skipped '{subject}': {e}")
                    break
    finally:
        driver.quit()

    if not results:
        print("No data was scraped -- nothing to save.")
        sys.exit(1)

    build_excel(results, OUTPUT_FILE)

if __name__ == "__main__":
    main()
